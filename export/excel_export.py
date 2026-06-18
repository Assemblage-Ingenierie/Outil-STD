"""Export Excel multi-onglets de toutes les tables de l'Outil STD."""
from __future__ import annotations
from io import BytesIO

import numpy as np
import pandas as pd


ROUGE = "E30513"
GRIS = "F2F2F2"


def _df_synthese(variantes, config, methode, dh_on, seuil_t1):
    rows = []
    for var in variantes:
        ind = var.indicateurs_batiment(config, methode)
        if dh_on:
            ind[f"DH > {seuil_t1:.0f}°C (°C·h)"] = var.dh_batiment(seuil_t1)
        rows.append({'Variante': var.nom, **ind})
    return pd.DataFrame(rows)


def _df_focus(variantes, config, methode, dh_on, seuil_t1, seuil_t2, zone):
    lib = "COCO" if methode == "coco" else "Givoni"
    rows = []
    for var in variantes:
        st_ = var.stats_temp(zone)
        syn = var.synthese_zone(zone)
        row = {
            'Variante': var.nom,
            'Surface (m²)': round(syn['surface_m2'], 0) if syn and not np.isnan(syn.get('surface_m2', float('nan'))) else np.nan,
            'Besoins ch. (kWh/m²)': round(syn['besoins_chaud_kwh_m2'], 1) if syn else np.nan,
            'Besoins fr. (kWh/m²)': round(syn['besoins_froid_kwh_m2'], 1) if syn else np.nan,
            'T min (°C)': round(st_['t_min'], 1),
            'T moy (°C)': round(st_['t_moy'], 1),
            'T max (°C)': round(st_['t_max'], 1),
            f'H > {seuil_t1}°C': var.heures_dessus_seuil(zone, seuil_t1),
            f'H > {seuil_t2}°C': var.heures_dessus_seuil(zone, seuil_t2),
            f'% hors {lib} 0 m/s': var.pct_hors_confort(zone, config, 0.0, methode),
            f'% hors {lib} 1 m/s': var.pct_hors_confort(zone, config, 1.0, methode),
        }
        if dh_on:
            row[f'DH > {seuil_t1:.0f}°C (°C·h)'] = var.degre_heures(zone, seuil_t1)
        row['Occupation (h/an)'] = var.heures_occupation(zone)
        row['Météo'] = var.meteo_affiche() or '—'
        rows.append(row)
    return pd.DataFrame(rows)


def _df_comparaison(var, config, methode, dh_on, seuil_t1, seuil_t2, zones):
    lib = "COCO" if methode == "coco" else "Givoni"
    rows = []
    for zone in zones:
        st_ = var.stats_temp(zone)
        syn = var.synthese_zone(zone)
        row = {
            'Zone': zone,
            'Surface (m²)': round(syn['surface_m2'], 0) if syn and not np.isnan(syn.get('surface_m2', float('nan'))) else np.nan,
            'Besoins ch. (kWh/m²)': round(syn['besoins_chaud_kwh_m2'], 1) if syn else np.nan,
            'Besoins fr. (kWh/m²)': round(syn['besoins_froid_kwh_m2'], 1) if syn else np.nan,
            'T min (°C)': round(st_['t_min'], 1),
            'T moy (°C)': round(st_['t_moy'], 1),
            'T max (°C)': round(st_['t_max'], 1),
            f'H > {seuil_t1}°C': var.heures_dessus_seuil(zone, seuil_t1),
            f'H > {seuil_t2}°C': var.heures_dessus_seuil(zone, seuil_t2),
            f'% hors {lib} 0 m/s': var.pct_hors_confort(zone, config, 0.0, methode),
            f'% hors {lib} 1 m/s': var.pct_hors_confort(zone, config, 1.0, methode),
        }
        if dh_on:
            row[f'DH > {seuil_t1:.0f}°C (°C·h)'] = var.degre_heures(zone, seuil_t1)
        rows.append(row)
    return pd.DataFrame(rows)


def _styler_entetes(writer):
    """Met en forme les en-têtes (fond rouge, texte blanc gras) sur chaque feuille."""
    from openpyxl.styles import PatternFill, Font, Alignment
    fill = PatternFill("solid", fgColor=ROUGE)
    font = Font(color="FFFFFF", bold=True)
    for ws in writer.book.worksheets:
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # largeur de colonnes auto (approximative)
        for col in ws.columns:
            longueur = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(longueur + 2, 10), 40)
        ws.freeze_panes = "A2"


def generer_excel(variantes, config, seuil_t1, seuil_t2, methode="givoni",
                  dh_on=False, zone_focus=None, var_comp=None, zones_comp=None,
                  df_recap=None, df_detail=None) -> BytesIO:
    """Construit un classeur Excel multi-onglets et renvoie un BytesIO."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # 1. Synthèse bâtiment
        _df_synthese(variantes, config, methode, dh_on, seuil_t1).to_excel(
            writer, sheet_name="Synthèse", index=False)

        # 2. Focus zone
        if zone_focus:
            _df_focus(variantes, config, methode, dh_on, seuil_t1, seuil_t2, zone_focus).to_excel(
                writer, sheet_name="Focus zone", index=False)

        # 3. Comparaison de zones
        if var_comp is not None and zones_comp:
            _df_comparaison(var_comp, config, methode, dh_on, seuil_t1, seuil_t2, zones_comp).to_excel(
                writer, sheet_name="Comparaison zones", index=False)

        # 4. Améliorations
        if df_recap is not None and not df_recap.empty:
            df_recap.reset_index().to_excel(writer, sheet_name="Améliorations (récap)", index=False)
        if df_detail is not None and not df_detail.empty:
            df_detail.to_excel(writer, sheet_name="Améliorations (descriptif)", index=False)

        _styler_entetes(writer)

    buf.seek(0)
    return buf
